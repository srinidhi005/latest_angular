#!/usr/bin/env python
# coding: utf-8

# In[59]:


import glob
import os
import openpyxl
import re
import sys
import warnings

warnings.filterwarnings(action="ignore")
import config
import pdftables_api
import mysql.connector
import datetime

API_KEY = '6yc5mavlgq4v'

units_pattern = config.units_pattern
company_names_pattern = config.company_names_pattern
statements = config.statements
sga_patterns = config.sga_patterns
cost_of_goods_sold = config.cost_of_goods_sold
total_revenue_patterns = config.total_revenue_patterns
product_dev = config.product_dev
sales_and_marketing = config.sales_and_marketing
total_operating_exp = config.total_operating_exp
net_interest_exp = config.net_interest_exp
other_income_exp = config.other_income_exp
tax_patterns = config.tax_patterns
statements_cash_flow = config.statements_cash_flow
depriciation_amort = config.depriciation_amort
gross_profit=config.gross_profit
ebt_patterns=config.ebt_patterns
ebit_patterns=config.ebit_patterns
net_income=config.net_income
r_and_d_pattern=config.r_and_d_pattern
in_process_pattern=config.in_process_pattern
# units_pattern = ["thousands","millions"] #acceptable units
# company_names_pattern = [".*Inc\.",".*CORPORATION",".*International",".*Subsidiaries",".*Incorporation"] #acceptable company names

### statements to consider
##statements = [
##".*Consolidated.*Statement.*Operations.*",
##".*Consolidated.*Statement.*Income.*",
##".*Consolidated.*Statement.*Earnings.*",
##]
##
##total_revenue = ["Total Net Revenue","Total Revenue","Total Revenues","Net Sale","Net Sales",
##                 "Revenue","Revenue","Net Revenue","Net Revenues","Sales to customer"]
##
##sga_patterns = ["General.*administrative.*","Selling.*informational.*administrative.*","Selling.*general.*administrative",
##                "Total.*selling.*administrative.*","Marketing.*administration.*research.*","Selling.*marketing.*and.*administrative.*expenses"]
##
##cost_of_goods_sold = ["Total costs and expenses","Total cost of revenues","Cost of Sale.*","COGs","Cost of goods sold","Cost of Revenue.?","Cost of products sold"]
##
##net_interest_exp = ["Interest expense.*income.*net.*","Interest income.*expense.*net.*","Interest, net","Interest expense.*net",
##                    "Interest expense","Interest and debt expense","Net interest expense"]
##other_income_exp = ["Other expenses","Other expense.*income.*net","Other expense.*net","Interest.*other.*income.*expense.*net","Other.*income/(loss).*net","Net.*other.*(income).*/.*expense","Other.*(income).*expense.*net","Other.*(income).*expense"]
##
##tax_patterns = [r"Income tax benefit \(provision\)","Income tax expense","Income tax expense (benefit)","Income tax provision","Provision/(benefit) for taxes on income","Provision for income taxes","Income taxes","Provision.*for.*taxes.*on.*income.*"]
##
##
##
##depriciation_amort=["Depreciation",
##"Amortization",
##"Depreciation and amortization",
##"Depreciation expense",
##"Depreciation and amortization of property, equipment and intangibles"
##"Depreciation, amortization, and other"
##"Depreciation and amortization expense"
##]
##
##statements_cash_flow=[".*Consolidated Statements of Cash Flows"]


def get_all_excel_files(pdf_folder_path):
    xlsx_files = glob.glob(pdf_folder_path + "/*.xlsx")  # collect all xlsx files
    return xlsx_files


def get_company_name_and_units_used(xlsx_files):
    """
    :param pdf_folder_path:
    :return: dictionary containing company and units used
    """
    found=0
    company_names = []
    units_used = []

    for xl_path in xlsx_files:  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        no_of_sheets = len(work_book.sheetnames)  # get sheets count
        for sheet_no in range(no_of_sheets):  # loop through all sheets
            work_sheet = work_book.worksheets[sheet_no]  # take sheet reference
            found = 0
            for j in range(1, 4):  # search upto 3rd row
                for i in range(1, 4):  # search upto 3rd column to get company name
                    if work_sheet.cell(row=j, column=i).value is not None:  # check non empty cells only
                        for com_name_pattern in company_names_pattern:  # loop through all possible copany names pattern
                            match_obj = re.match(com_name_pattern, str(work_sheet.cell(row=j, column=i).value),
                                                 re.I)  # match with ignoring case
                            if match_obj:
                                company_names.append(match_obj.group(0))
                                found = 1
                                break

                    if found == 1:
                        break
            if found == 1:
                break
        if found != 1:
            company_names.append("NA")
    # for company in company_names:
    #     print(company)
    file_names = []
    for xl_path in xlsx_files:  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        no_of_sheets = len(work_book.sheetnames)  # get sheets count
        for sheet_no in range(no_of_sheets):  # loop through all sheets
            work_sheet = work_book.worksheets[sheet_no]  # take sheet reference
            found = 0
            for j in range(1, 7):  # search upto 6th row
                for i in range(1, 4):  # search upto 3rd column to get units used
                    if work_sheet.cell(row=j, column=i).value is not None:
                        for unit_pattern in units_pattern:
                            match_obj = re.search(unit_pattern, str(work_sheet.cell(row=j, column=i).value), re.I)
                            if match_obj:
                                units_used.append(match_obj.group(0))
                                found = 1
                                break
                    if found == 1:
                        break
            if found == 1:
                break
        file_names.append(filename)
    return dict(zip(company_names, units_used)), dict(zip(file_names, company_names))


def get_required_sheets(xlsx_files):
    """

    :return: returns sheet references
    """
    sheet_ref = {}
    for xl_path in xlsx_files:  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        # print(filename)
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        no_of_sheets = len(work_book.sheetnames)  # get sheets count
        for sheet_no in range(no_of_sheets):  # loop through all sheets
            work_sheet = work_book.worksheets[sheet_no]  # take sheet reference
            # print(work_sheet.title,"in this page")
            found = 0
            for j in range(1, 8):  # search upto 3rd row
                for i in range(1, 5):  # search upto 3rd column to get statements name
                    if work_sheet.cell(row=j, column=i).value is not None:
                        for statement in statements:
                            # print(statement)
                            match_obj = re.match(statement, str(work_sheet.cell(row=j, column=i).value),
                                                 re.I | re.MULTILINE)
                            if match_obj:
                                # print(work_sheet.title)
                                sheet_ref[xl_path] = work_sheet.title
                                found = 1
                                break
                if found == 1:
                    break
            if found == 1:
                break

    return sheet_ref


def check_multi_cell_and_split(multiline_str):
    # print("multiline cell function called")
    terms = re.findall("[A-Z][^A-Z]*", multiline_str)
    return terms


def get_years(xlsx_files, filename_company):
    """
    fetch years and retueen comapny  year dict
    :return:
    """
    years_dict = {}

    for xl_path, page_ref in xlsx_files.items():  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        work_sheet = work_book[page_ref]  # take sheet reference
        flg = 1
        count = 0
        years = []
        for j in range(1, 10):

            years_split = []
            for i in range(1, 10):
                if work_sheet.cell(row=j, column=i).value is not None:
                    ##                    years_split=str.split(work_sheet.cell(row=j, column=i).value)
                    ##                    print(years_split)
                    match_obj = re.match("^2\d{3}$", str(work_sheet.cell(row=j, column=i).value), re.I)

                    if match_obj:
                        ##                        print(work_sheet.cell(row=j, column=i).value)
                        ##                        print("hi")
                        count += 1
                        # print(match_obj.group(0),"year")
                        years.append(match_obj.group(0))
                    if count == 3:
                        flg = 0
                        break
            years.sort(reverse = True)
            years_dict[filename_company[filename]] = years
            if flg == 0:
                break
    return years_dict


def get_required_values(xlsx_files, search_term, filename_company):
    """
    :return: returns sga and total revenue

    apple is having empty. its avilable in next
    """
    found=0
    term_values = []
    term_patterns = search_term
    term_dict = {}
    for xl_path, page_ref in xlsx_files.items():  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        # print(filename)
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        work_sheet = work_book[page_ref]  # take sheet reference
        found = 0
        row_count = work_sheet.max_row
        column_count = work_sheet.max_column

        for j in range(1, row_count):  # search upto 25th row
            for i in range(1, column_count):  # search upto 3rd column to get units used
                if work_sheet.cell(row=j, column=i).value is not None:
                    for term in term_patterns:
                        match_obj = re.search(term, str(work_sheet.cell(row=j, column=i).value), re.I)
                        if match_obj:
                            # print(match_obj)
                            # print(str(work_sheet.cell(row=j, column=i).value))

                            terms_found = check_multi_cell_and_split(work_sheet.cell(row=j, column=i).value)
                            if len(terms_found) == 1:
                                if re.match(term, terms_found[0], re.I):
                                    for k in range(10):
                                        sga_match_obj = re.search("\d+[\.\d+].*",
                                                                  str(work_sheet.cell(row=j, column=k + 2).value))

                                        if sga_match_obj:
                                            # print(str(work_sheet.cell(row=j, column=k+2).value))
                                            term_values.append(str(work_sheet.cell(row=j, column=k + 2).value))
                                            found = 1
                                            term_dict[filename_company[filename]] = term_values
                                        elif work_sheet.cell(row=j, column=k + 2).value is not None and not "$":
                                            term_values.append(str(0))
                                            found = 1
                                        elif work_sheet.cell(row=j, column=k+2).value =='-':
                                            term_values.append(str(0))
                                            found=1
                                    if found == 1:
                                        break


                            elif len(terms_found) > 1:
                                print("Multiline found", terms_found)
                                for k in range(10):
                                    no_of_values = re.findall("[1-9,.]+",
                                                              str(work_sheet.cell(row=j, column=k + 2).value))
                                    if no_of_values != []:
                                        # print(str(work_sheet.cell(row=j, column=k+2).value))

                                        for key, value in zip(terms_found, no_of_values):
                                            if re.match(term, key, re.I):
                                                term_values.append(value)
                                                break

                                        found = 1
                                term_dict[filename_company[filename]] = term_values
                                if found == 1:
                                    break

                        if found == 1:
                            break

        if found != 1:
            for i in range(len((list(company_years)[0]))):
                term_values.append(str('0'))
                term_dict[filename_company[filename]] = term_values
    return term_dict



def get_income_values(xlsx_files, search_term, filename_company):
    term_values = []
    term_patterns = search_term
    term_dict = {}
    row_col_list=[]
    for xl_path, page_ref in xlsx_files.items():  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        work_sheet = work_book[page_ref]  # take sheet reference
        found = 0
        row_count = work_sheet.max_row
        column_count = work_sheet.max_column

        for j in range(1, row_count):  # search upto 25th row
            for i in range(1, column_count):  # search upto 3rd column to get units used
                if work_sheet.cell(row=j, column=i).value is not None:
                    for term in term_patterns:
                        match_obj = re.search(term, str(work_sheet.cell(row=j, column=i).value), re.I|re.M)
                        
                        if match_obj: 
                            terms_found = check_multi_cell_and_split(work_sheet.cell(row=j, column=i).value)
                            rowid_match=j-1
                            columnid_match=i
                            row_col_list.append(rowid_match)
                            if re.match(term, terms_found[0], re.I):
                                for k in range(10):
                                    ni_match_obj = re.search("\d+[\.\d+].*",
                                                              str(work_sheet.cell(row=j, column=k + 2).value))
            
                                    if ni_match_obj:
                                        #print(str(work_sheet.cell(row=j, column=k+2).value))
                                        term_values.append(str(work_sheet.cell(row=j, column=k + 2).value))
                                        found = 1
                                        term_dict[filename_company[filename]] = term_values
                                        
                                    elif work_sheet.cell(row=j, column=k+2).value is not None and not "$":
                                        term_values.append(str(0))
                                        found=1       
                                if found == 1:
                                        break
                    

                        if found == 1:
                            break
                        
    if found!=1:
        rowid_match=row_col_list[0]
        if match_obj is None:
            for k in range(10):
                ni_match_obj = re.search("\d+[\.\d+].*",str(work_sheet.cell(row=rowid_match, column=k + 2).value))               
                if ni_match_obj:
                    term_values.append(str(work_sheet.cell(row=rowid_match, column=k + 2).value))
                    term_dict[filename_company[filename]] = term_values                  
                elif work_sheet.cell(row=rowid_match, column=k+2).value is not None and not "$":
                    term_values.append(str(0))  
                    term_dict[filename_company[filename]] = term_values
    return term_dict    






def get_required_gp_values(xlsx_files, search_term, filename_company):
    """
    :return: returns sga and total revenue

    apple is having empty. its avilable in next
    """
    term_values = []
    term_patterns = search_term
    term_dict = {}
    for xl_path, page_ref in xlsx_files.items():  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        # print(filename)
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        work_sheet = work_book[page_ref]  # take sheet reference
        found = 0
        row_count = work_sheet.max_row
        column_count = work_sheet.max_column

        for j in range(1, row_count):  # search upto 25th row
            for i in range(1, column_count):  # search upto 3rd column to get units used
                if work_sheet.cell(row=j, column=i).value is not None:
                    for term in term_patterns:
                        match_obj = re.search(term, str(work_sheet.cell(row=j, column=i).value), re.I|re.M)
                        
                        if match_obj:
                            #print(match_obj)
                           
                            
                            terms_found = check_multi_cell_and_split(work_sheet.cell(row=j, column=i).value)
                            if len(terms_found) == 1:
                                if re.match(term, terms_found[0], re.I):
                                    for k in range(10):
                                        sga_match_obj = re.search("\d+[\.\d+].*",str(work_sheet.cell(row=j, column=k + 2).value)) 
                                        if sga_match_obj:
                                            #print(str(work_sheet.cell(row=j, column=k+2).value))
                                            term_values.append(str(work_sheet.cell(row=j, column=k + 2).value))
                                            found = 1
                                            term_dict[filename_company[filename]] = term_values
                                            
                                        elif work_sheet.cell(row=j, column=k+2).value is not None and not "$":
                                            term_values.append(str(0))
                                            found=1       
                                    if found == 1:
                                            break
                            

                            elif len(terms_found) > 1:
                                print("Multiline found", terms_found)
                                for k in range(10):
                                    no_of_values = re.findall("[1-9,.]+",
                                                              str(work_sheet.cell(row=j, column=k + 2).value))
                                    if no_of_values != []:
                                        # print(str(work_sheet.cell(row=j, column=k+2).value))

                                        for key, value in zip(terms_found, no_of_values):
                                            if re.match(term, key, re.I):
                                                term_values.append(value)
                                                break

                                        found = 1
                                term_dict[filename_company[filename]] = term_values
                                if found == 1:
                                    break

                        if found == 1:
                            break
                        
    return term_dict




def convert_to_thousand(terms, unit_dict):
    """
    if values are in terms of Million then Multiply by 1000
    to make it in thousands
    :return:
    """
    temp_dict = {}
    for key, value in terms.items():
        modified_value = []
        for val in value:
            if re.search("\(", val):
                val = val.strip("(")  # repmove (
                val = val.strip(")")  # remove )
                val = val.replace(",", "")  # remove comma
                val = -float(val)  # conver to float and minus it
            else:
                val = val.strip("$")
                val = val.replace(",", "")  # remove comma
                val = float(val)  # conver to float and minus it
            modified_value.append(val)
        if key in unit_dict.keys():
            if unit_dict[key].lower() == "millions":
                modified_value = [val * 1000 for val in modified_value]
            else:
                modified_value = modified_value

        temp_dict[key] = modified_value
    return temp_dict


def get_required_sheets_cash_flows(xlsx_files):
    """

    :return: returns sheet references
    """
    sheet_ref = {}
    for xl_path in xlsx_files:  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        # print(filename)
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        no_of_sheets = len(work_book.sheetnames)  # get sheets count
        for sheet_no in range(no_of_sheets):  # loop through all sheets
            work_sheet = work_book.worksheets[sheet_no]  # take sheet reference
            # print(work_sheet.title,"in this page")
            found = 0
            for j in range(1, 8):  # search upto 3rd row
                for i in range(1, 5):  # search upto 3rd column to get statements name
                    if work_sheet.cell(row=j, column=i).value is not None:
                        for statement_cash in statements_cash_flow:
                            # print(statement_cash)
                            match_obj = re.match(statement_cash, str(work_sheet.cell(row=j, column=i).value), re.I)
                            if match_obj:
                                # print(work_sheet.title)
                                sheet_ref[xl_path] = work_sheet.title
                                found = 1
                                break
                if found == 1:
                    break
            if found == 1:
                break

    return sheet_ref


def get_depriciation_and_amortizatoin(xlsx_files):
    """
    :return: returns tax

    """
    dep_amort_val_dict = {}
    dep_amort_val = []
    for xl_path, page_ref in xlsx_files.items():  # loop through all xlsx files
        path, filename = os.path.split(xl_path)  # split file and path
        # print(xl_path,page_ref)
        work_book = openpyxl.load_workbook(xl_path)  # load workbook
        work_sheet = work_book[page_ref]  # take sheet reference
        found = 0
        row_count = work_sheet.max_row
        column_count = work_sheet.max_column
        for j in range(1, row_count):  # search upto 25th row
            for i in range(1, column_count):  # search upto 3rd column to get units used
                if work_sheet.cell(row=j, column=i).value is not None:

                    for dep_amort in depriciation_amort:
                        match_obj = re.search(dep_amort, str(work_sheet.cell(row=j, column=i).value), re.I)
                        print(match_obj)
                        # print(str(work_sheet.cell(row=j, column=i).value))
                        # print(taxes)
                        if match_obj:
                            # print("tax_vals=>",match_obj)
                            dep_amort_val = []
                            for k in range(10):
                                match_obj_dep_amort_val = re.match("\(?\d+(,\d+)*(\.\d+)?\)?",
                                                                   str(work_sheet.cell(row=j, column=k + 2).value))
                                # re.match("\d+[,\.\d+].*",str(work_sheet.cell(row=j, column=k+2).value))  or re.match("\d+[,\.\d+].*",str(work_sheet.cell(row=j, column=k+2).value))

                                # print(match_obj_tax_val)
                                if match_obj_dep_amort_val:
                                    # print(str(work_sheet.cell(row=j, column=k+2).value))
                                    dep_amort_val.append(match_obj_dep_amort_val.group(0))
                                    dep_amort_val_dict[xl_path] = dep_amort_val
                                    # print(cogs_val)
                                    # print(cogs_val_dict)
                                    found = 1

                                    # break;

                            if found == 1:
                                break
                        if found == 1:
                            break

    return dep_amort_val_dict


def convert_pdf_to_excel(pdf_path):
    c = pdftables_api.Client(API_KEY)

    path, file = os.path.split(pdf_path)
    excelfile = file.split(".")[-2]
    excelfile_path = "/home/srinidhi/angular/extractor/temp"
    c.xlsx(pdf_path, excelfile_path + "/temporary")
    print(excelfile_path,"File Created")
    return excelfile_path


if __name__ == "__main__":
    op = open("/home/srinidhi/angular/extractor/Extraction_logger.txt", "a")
    created_on = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        pdf_path = sys.argv[1]
        company_name = sys.argv[2]
        period = sys.argv[3]
        user = sys.argv[4]
        path, filename = os.path.split(pdf_path)
        filename = filename
        industry = sys.argv[5]
        statement_type = sys.argv[6]

        # pdf_file_path = r"D:\original_xls_bkp"  # excel files path
        pdf_file_path = pdf_path
        excel_files = convert_pdf_to_excel(pdf_file_path)
        excel_files = get_all_excel_files(excel_files)
        company_and_unit, filename_company = get_company_name_and_units_used(excel_files)
        print("COMPANY Units", company_and_unit)
        company = list(company_and_unit.keys())[0]
        sheet_ref = get_required_sheets(excel_files)
        
        company_years = get_years(sheet_ref, filename_company)
        print("Company Years", len(company_years), company_years)
        company_years = company_years.values()
        print(company_years,list(company_years))
        


        
#        sga = get_required_values(sheet_ref, sga_patterns, filename_company)
#        sga = convert_to_thousand(sga, company_and_unit)
#        sga = sga.values()
#        print("SGA", len(sga), sga)

        total_revenue = get_required_values(sheet_ref, total_revenue_patterns, filename_company)
        total_revenue = convert_to_thousand(total_revenue, company_and_unit)
        total_revenue = total_revenue.values()
        print("Revenue", len(total_revenue), total_revenue)
        
        gross_profit_values={"a": []}
        gross_profit_values=get_required_gp_values(sheet_ref, gross_profit, filename_company)
        gross_profit_values = convert_to_thousand(gross_profit_values, company_and_unit)
        gross_profit_values=gross_profit_values.values()
        print("Gross Profit",gross_profit_values)

        net_interest_exp = get_required_values(sheet_ref, net_interest_exp, filename_company)
        net_interest_exp = convert_to_thousand(net_interest_exp, company_and_unit)
        net_interest_exp = net_interest_exp.values()
        print("Net Interest Expense", len(net_interest_exp), net_interest_exp)

        other_income_exp = get_required_values(sheet_ref, other_income_exp, filename_company)
        other_income_exp = convert_to_thousand(other_income_exp, company_and_unit)
        other_income_exp = other_income_exp.values()
        print("Other Income Expense", len(other_income_exp), other_income_exp)

        tax_val_dict = get_required_values(sheet_ref, tax_patterns, filename_company)
        tax_val_dict = convert_to_thousand(tax_val_dict, company_and_unit)
        tax_val_dict = tax_val_dict.values()
        print("Taxes", len(tax_val_dict), tax_val_dict)

        ebit = get_required_values(sheet_ref, ebit_patterns, filename_company)
        ebit = convert_to_thousand(ebit, company_and_unit)
        ebit = ebit.values()
        print("ebit", len(ebit), ebit)


        
        if all(v == 0 for v in list(ebit)[0]):
            sga = get_required_values(sheet_ref, sga_patterns, filename_company)
            sga = convert_to_thousand(sga, company_and_unit)
            sga = sga.values()
            r_and_d=get_required_values(sheet_ref, r_and_d_pattern, filename_company)
            r_and_d = convert_to_thousand(r_and_d, company_and_unit)
            r_and_d = r_and_d.values()
            in_process=get_required_values(sheet_ref, in_process_pattern, filename_company)
            in_process = convert_to_thousand(in_process, company_and_unit)
            in_process = in_process.values()
            
            sga_rd={"a": []}
            for i in range(len((list(company_years)[0]))):
                sg_rd=list(sga)[0][i] + list(r_and_d)[0][i] +list(in_process)[0][i]
                sga_rd["a"].append(sg_rd)
            sga=sga_rd.values()
            print("sga_final",sga) 


            ebit_dict={"a": []}
            for i in range(len((list(company_years)[0]))):
                
                EBIT=list(gross_profit_values)[0][i]-list(sga)[0][i]
                ebit_dict["a"].append(EBIT)
            ebit=ebit_dict.values()
                 
        cogs = get_required_values(sheet_ref, cost_of_goods_sold, filename_company)
        cogs = convert_to_thousand(cogs, company_and_unit)
        cogs = cogs.values()
        print("COGS", len(cogs), cogs)

        net_income = get_income_values(sheet_ref, net_income, filename_company)
        net_income = convert_to_thousand(net_income, company_and_unit)
        net_income = net_income.values()
        print("net_income", len(net_income), net_income)


        ebt = get_required_values(sheet_ref, ebt_patterns, filename_company)
        ebt = convert_to_thousand(ebt, company_and_unit)
        ebt = ebt.values()
        print("ebt", len(ebt), ebt)
        
        sh_ref = get_required_sheets_cash_flows(excel_files)
        dep_amort = get_required_values(sh_ref, depriciation_amort, filename_company)
        dep_amort = convert_to_thousand(dep_amort, company_and_unit)
        dep_amort = dep_amort.values()
        print("statements_cash_flow", len(dep_amort), dep_amort)

        if not gross_profit_values:
            gross_profit={"a": []}
            for i in range(len((list(company_years)[0]))):
                gp = list(total_revenue)[0][i] - abs(list(cogs)[0][i])
                gross_profit["a"].append(gp)
            gross_profit_values=gross_profit.values()
        print("gross_profit_values",gross_profit_values)

        connection = mysql.connector.connect(host="35.225.71.54",
                                             database='finance',
                                             user='investor',
                                             password='investor')
        if connection.is_connected():
            db_Info = connection.get_server_info()
            print("Connected to MySQL Server version ", db_Info)
            cursor = connection.cursor()

        projections = 4  # hardcodedd

        query = "insert into company_master (companyname,company,period,actuals,projections,createdby,createdon,filename,industry,statementtype)" \
                "values ('" + company_name + "','" + company + "','" + period + "'," + str(len(list(company_years)[0])) + "," + str(projections) + ",'" + user + "','" + created_on + "','" + filename + "','" + industry + "','" + statement_type + "')"
        print(query,"===")
        cursor.execute(query)
        connection.commit()
        latest = 0
        for i in range(len((list(company_years)[0]))):
            sga=list(gross_profit_values)[0][i]-list(ebit)[0][i]
            print(list(gross_profit_values)[0][i])
            net_interest_exp=list(ebit)[0][i]-list(ebt)[0][i]
            taxes=list(ebt)[0][i]-list(net_income)[0][i]
            ebitda = list(ebit)[0][i] + list(dep_amort)[0][i]
            grossprofitmargin = float((list(gross_profit_values)[0][i] / list(total_revenue)[0][i]) * 100)      
            ebitmargin = float((list(ebit)[0][i] / list(total_revenue)[0][i]) * 100)
            ebitdamargin = float((ebitda/ list(total_revenue)[0][i]) * 100)
            ebtmargin = float((list(ebt)[0][i] / list(total_revenue)[0][i]) * 100)
            netincomemargin = float((list(net_income)[0][i] /list(total_revenue)[0][i]) * 100)
            query = "insert into company_actuals (companyname,asof,latest,totalrevenue,cogs,sga,da,netinterest,otherincome," \
                    "taxes,grossprofit,ebit,ebitda,netincome,grossprofitmargin,ebitmargin,ebitdamargin,ebtmargin,netincomemargin,ebt) values(" \
                    "'" + company_name + "'," +str(list(company_years)[0][i]) + "," + str(
                        latest) + "," + str(list(total_revenue)[0][i]) + "," + str(list(cogs)[0][i]) + "," + str(sga) + "," + str(
                        list(dep_amort)[0][i]) + "," + str(net_interest_exp) + "," + str(list(other_income_exp)[0][i]) + "," + str(taxes) + "," + str(
                        list(gross_profit_values)[0][i]) + "," + str(list(ebit)[0][i]) + "," + str(ebitda) + "," + str(list(net_income)[0][i]) + "," + str(
                        grossprofitmargin) + "," + str(ebitmargin) + "," + str(ebitdamargin) + "," + str(
                        ebtmargin) + "," + str(netincomemargin) + "," + str(list(ebt)[0][i]) + ")"                

            cursor.execute(query)
            connection.commit()  # save records
            latest -= 1
    except Exception as e:
        print("Error reading data from MySQL table", e)
        op.writelines(str(created_on)+"\n\n")
        op.writelines(str(e))
        op.writelines("\n\n")
